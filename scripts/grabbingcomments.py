from openpyxl import load_workbook
import os

file_path = r'C:\Users\corey\github\models\CTVA.xlsx'

# Check if the file exists
if not os.path.exists(file_path):
    print(f"Oops! The file at {file_path} does not exist.")
else:
    workbook = load_workbook(filename=file_path)

    if 'model' in workbook.sheetnames:
        # Select the 'model' worksheet
        worksheet = workbook['model']

        # Iterate through the cells to find comments
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment:
                    print(f"Cell {cell.coordinate} has comment: {cell.comment.text}")
    else:
        print("The sheet named 'model' doesn't exist in this workbook.")
